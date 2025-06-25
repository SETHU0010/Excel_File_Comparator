import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def compare_excels(file1, file2, key_col=None, highlight_options=None):
    try:
        df1 = pd.read_excel(file1)
        df2 = pd.read_excel(file2)

        if list(df1.columns) != list(df2.columns):
            return None, "The Excel files have different columns. Please upload files with matching structure.", {}

        df1.fillna("NaN", inplace=True)
        df2.fillna("NaN", inplace=True)

        max_rows = max(len(df1), len(df2))
        df1 = df1.reindex(range(max_rows)).fillna("NaN")
        df2 = df2.reindex(range(max_rows)).fillna("NaN")

        diff_df = df2.copy()
        output = BytesIO()
        diff_df.to_excel(output, index=False)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        # Colors
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Modified
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # New row
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Deleted row

        # Stats
        new_rows, deleted_rows, modified_cells = 0, 0, 0

        for row in range(max_rows):
            row1 = df1.iloc[row]
            row2 = df2.iloc[row]

            if row1.equals(row2):
                continue
            elif row1.isnull().all():
                new_rows += 1
                if highlight_options.get("highlight_new"):
                    for col in range(len(df2.columns)):
                        ws.cell(row=row + 2, column=col + 1).fill = green_fill
            elif row2.isnull().all():
                deleted_rows += 1
                if highlight_options.get("highlight_deleted"):
                    for col in range(len(df1.columns)):
                        ws.cell(row=row + 2, column=col + 1).value = df1.iloc[row, col]
                        ws.cell(row=row + 2, column=col + 1).fill = red_fill
            else:
                for col in range(len(df1.columns)):
                    val1 = df1.iloc[row, col]
                    val2 = df2.iloc[row, col]
                    if val1 != val2:
                        modified_cells += 1
                        if highlight_options.get("highlight_changes"):
                            ws.cell(row=row + 2, column=col + 1).fill = yellow_fill

        output_final = BytesIO()
        wb.save(output_final)
        output_final.seek(0)

        stats = {
            "Total Rows Compared": max_rows,
            "New Rows": new_rows,
            "Deleted Rows": deleted_rows,
            "Modified Cells": modified_cells
        }

        return output_final, None, stats

    except Exception as e:
        return None, str(e), {}


# ---------------- Streamlit UI ---------------- #

st.set_page_config(page_title="Excel Comparator", page_icon="📊")
st.title("📊 Excel File Comparator")
st.markdown("Compare two Excel files and highlight the differences.")

with st.expander("📁 Upload Excel Files"):
    file1 = st.file_uploader("Upload First Excel File", type=["xlsx", "xls"], key="file1")
    file2 = st.file_uploader("Upload Second Excel File", type=["xlsx", "xls"], key="file2")

if file1 and file2:
    df1_preview = pd.read_excel(file1, nrows=5)
    st.markdown("✅ Sample Data from First File")
    st.dataframe(df1_preview)

    df2_preview = pd.read_excel(file2, nrows=5)
    st.markdown("✅ Sample Data from Second File")
    st.dataframe(df2_preview)

    columns = list(df1_preview.columns)
    st.markdown("🔑 Choose Matching Options")
    key_col = st.selectbox("Select Key Column (optional)", options=[""] + columns)

    st.markdown("🎨 Highlight Options")
    col1, col2, col3 = st.columns(3)
    highlight_new = col1.checkbox("Highlight New Rows", value=True)
    highlight_deleted = col2.checkbox("Highlight Deleted Rows", value=True)
    highlight_changes = col3.checkbox("Highlight Modified Cells", value=True)

    if st.button("🔍 Compare Files"):
        with st.spinner("Comparing Excel files..."):
            result_file, error, stats = compare_excels(
                file1, file2, key_col if key_col else None,
                highlight_options={
                    "highlight_new": highlight_new,
                    "highlight_deleted": highlight_deleted,
                    "highlight_changes": highlight_changes
                }
            )

        if error:
            st.error(f"❌ Error: {error}")
        else:
            st.success("✅ Comparison completed successfully!")
            st.markdown("📈 **Summary Report:**")
            st.markdown(f"""
            - 🔢 Total Rows Compared: {stats['Total Rows Compared']}
            - 🟩 New Rows: {stats['New Rows']}
            - 🟥 Deleted Rows: {stats['Deleted Rows']}
            - 🟨 Modified Cells: {stats['Modified Cells']}
            """)
            st.download_button(
                label="📥 Download Highlighted Comparison Report",
                data=result_file,
                file_name="Comparison_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
